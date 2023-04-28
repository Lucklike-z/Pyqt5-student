import os
import sys

import openpyxl
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *

import config
import db
from ui.home import Ui_Form_home
from ui.home_add import Ui_home_add
from ui.home_alter import Ui_home_alter
from ui.home_del_user import Ui_Form_del_user
from ui.home_query import Ui_Form_Query
from ui.home_regedit import Ui_Form_regedit
from ui.home_repass import Ui_Form_repass
from ui.login import Ui_Form_Login


class MyWindow(QWidget, Ui_Form_Login):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.setupUi(self)
        self.signal()
        self.check_mysql()

    def signal(self):
        self.setWindowIcon(QIcon('./ui/icon.png'))
        self.pushButton_login.clicked.connect(self.on_login)
        self.lineEdit_pass.returnPressed.connect(self.on_login)
        # self.mysql_connect.create_user()

    def on_login(self):
        self.winman = Winmain()
        self.username = self.lineEdit_user.text().strip()
        self.password = self.lineEdit_pass.text().strip()
        if self.mysql_connect.query(self.username, self.password):
            QMessageBox.information(self, '提示', '登录成功')
            config.login_username = self.username
            self.winman.ui_user(config.login_username)
            self.winman.show()
            self.hide()
            return
        QMessageBox.warning(self, '提示', '用户名或密码错误')
        # print(self.data())

    def check_mysql(self):
        self.mysql_connect = db.db_mysql()
        if self.mysql_connect.mysql_connect():
            if self.mysql_connect.create_mysql():
                if self.mysql_connect.show_user() == False:
                    self.mysql_connect.create_user()
                    return
                else:
                    return
            QMessageBox.warning(self, '连接出错', '数据库表创建失败，请删除数据库重试')
            sys.exit()
        else:
            QMessageBox.warning(self, '连接出错', '数据库连接失败，请检查数据库配置')
            sys.exit()


class Winmain(QWidget, Ui_Form_home):
    def __init__(self):
        super(Winmain, self).__init__()
        self.setupUi(self)
        self.WinMain()
        self.table_ui()
        self.show_student()
        self.tableWidget.verticalHeader().setHidden(True)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def WinMain(self):
        self.pushButton_redata.clicked.connect(self.show_student)
        self.pushButton_add.clicked.connect(self.add_sutdnet)
        self.tableWidget.cellPressed.connect(self.row_table_data)
        self.pushButton_del.clicked.connect(self.del_table_data)
        self.pushButton_admin_out.clicked.connect(self.out_admin)
        self.pushButton_readd.clicked.connect(self.revise_data)
        self.pushButton_query.clicked.connect(self.query_data)
        self.pushButton_load.clicked.connect(self.excel_data_load)
        self.pushButton_out.clicked.connect(self.excel_data_save)
        self.pushButton_regedit.clicked.connect(self.regedit)
        self.pushButton_repass.clicked.connect(self.repass)
        self.pushButton_del_user.clicked.connect(self.del_user)

    def regedit(self):
        self.regedit_user = regedit()
        self.regedit_user.show()

    def repass(self):
        self.repass_user = repass()
        self.repass_user.show()

    def del_user(self):
        self.del_users = del_class()
        self.del_users.show()

    def out_admin(self):
        self.login = MyWindow()
        choice = QMessageBox.question(
            self,
            '确认',
            '确定要退出{}吗？'.format(config.login_username))

        if choice == QMessageBox.Yes:
            self.login.show()
            self.close()
            return
        if choice == QMessageBox.No:
            return

    def ui_user(self, admin):
        self.label_admin.setText(admin)
        self.label_admin.setStyleSheet("color: red;")

    def show_student(self):
        config.result.clear()
        self.tableWidget.setRowCount(0)
        self.tableWidget.clearContents()

        self.tableWidget.setColumnCount(6)
        rowcount = self.tableWidget.rowCount()
        self.mysql_connect = db.db_mysql()
        self.mysql_connect.query_data()
        for s in config.result:
            self.tableWidget.insertRow(rowcount)
            col = 0
            for key, value in s.items():
                item = QTableWidgetItem(value)
                self.tableWidget.setItem(rowcount, col, item)
                col += 1

    def table_ui(self):
        self.tableWidget.setSelectionMode(QTableWidget.SingleSelection)
        self.tableWidget.setShowGrid(True)
        self.tableWidget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.tableWidget.setSelectionBehavior(QTableWidget.SelectRows)
        self.tableWidget.setSelectionMode(QTableWidget.SingleSelection)

    def row_table_data(self, row):
        config.row_data.clear()
        num_cols = self.tableWidget.columnCount()
        for col in range(num_cols):
            # 获取该单元格的文本内容
            item = self.tableWidget.item(row, col)
            # 将该单元格的文本内容添加到行数据列表中
            config.row_data.append(item.text())
        print(config.row_data)

    def del_table_data(self):
        if self.mysql_connect.delete_data(config.row_data):
            QMessageBox.information(self, '提示', '删除成功')
            self.show_student()
            return
        if self.mysql_connect.delete_data(config.row_data) == None:
            QMessageBox.warning(self, '提示', '你要删除什么呢？')
            return
        QMessageBox.warning(self, '提示', '删除失败')

    def query_data(self):
        self.query = query_data()
        self.query.show()

    def add_sutdnet(self):
        self.add = home_add()
        self.add.show()

    def revise_data(self):
        self.win_alter = home_alter()
        try:
            self.win_alter.lineEdit_id.setText(config.row_data[0])
            self.win_alter.lineEdit_id.setFocusPolicy(Qt.NoFocus)  # 设置不可编辑
            self.win_alter.lineEdit_name.setText(config.row_data[1])
            self.win_alter.lineEdit_class.setText(config.row_data[2])
            self.win_alter.lineEdit_yuwen.setText(config.row_data[3])
            self.win_alter.lineEdit_math.setText(config.row_data[4])
            self.win_alter.lineEdit_english.setText(config.row_data[5])
        except:
            QMessageBox.warning(self, '提示', '你要修改什么呢？')
            self.win_alter.close()
            return
        self.win_alter.show()

    def excel_data_load(self):
        file_path, _ = QFileDialog.getOpenFileName(None, "选择文件", "", "Excel Files (*.xlsx)")
        # 输出文件的绝对路径
        filepath, _filename = os.path.split(file_path)
        filename, extension = os.path.splitext(_filename)
        print('文件路径', filepath, '文件名', _filename, '文件类型', extension)
        if extension == '.xlsx':
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            # 定义要读取的单元格范围
            cell_range = worksheet['A2:F{}'.format(worksheet.max_row)]
            # 迭代指定范围内的单元格，将每个单元格的值添加到列表中
            config.cell_values.clear()
            for row in cell_range:
                row_values = []
                for cell in row:
                    row_values.append(cell.value)
                config.cell_values.append(row_values)
            # 输出读取的值
            print(config.cell_values)
            if config.cell_values != []:
                choice = QMessageBox.question(
                    self, '确认',
                    '确定要继续吗？')

                if choice == QMessageBox.Yes:
                    print(config.cell_values)
                    if self.mysql_connect.xlsx_load():
                        QMessageBox.information(self, '提示', '数据导入成功')
                        return
                    if self.mysql_connect.xlsx_load() == None:
                        QMessageBox.warning(self, '提示', '导入数据的学号重复，请检查')
                    else:
                        QMessageBox.critical(self, '提示', '数据导入失败')
                    return
                if choice == QMessageBox.No:
                    return
            QMessageBox.critical(self, '提示', '读取文件错误')
        return

    def excel_data_save(self):
        choice = QMessageBox.question(
            self, '确认',
            '确定要导出吗？')

        if choice == QMessageBox.Yes:
            if self.mysql_connect.xlsx_save() == False:
                QMessageBox.critical(self, '提示', '导出失败')
                return
            # 保存到一个excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            # 写入表头
            ws.append(['学号', '姓名', '班级', '语文', '数学', '英语'])
            # 执行查询操作
            row = self.mysql_connect.xlsx_save()
            # 写入数据
            for row in row:
                ws.append([row[0], row[1], row[2], row[3], row[4], row[5]])
            # 弹出打开或保存文件对话框
            file_name, _ = QFileDialog.getSaveFileName(None, "Save File", "", "Excel Files (*.xlsx)")

            # 如果用户选择了文件，那么打印文件名和路径
            if file_name:
                wb.save(file_name)
                print(f"选择的文件名和路径为：{file_name}")
                QMessageBox.information(self, '提示', '导出成功')
                return
            return
        if choice == QMessageBox.No:
            return


class home_add(QWidget, Ui_home_add):
    def __init__(self):
        super(home_add, self).__init__()
        self.setupUi(self)
        self.home()

    def home(self):
        self.pushButton_add.clicked.connect(self.data_add)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def data_add(self):
        self.home_1 = Winmain()
        self.mysql_connect = db.db_mysql()
        config.Student_data.clear()
        sid = self.lineEdit_id.text().strip()
        sname = self.lineEdit_name.text().strip()
        sclass = self.lineEdit_class.text().strip()
        smath = self.lineEdit_math.text().strip()
        syuwen = self.lineEdit_yuwen.text().strip()
        senglish = self.lineEdit_english.text().strip()
        student_data = {'id': str(sid), 'name': str(sname), 'class': str(sclass), 'math': str(smath),
                        'yuwen': str(syuwen), 'english': str(senglish)}
        config.Student_data.append(student_data)
        print(config.Student_data)
        try:
            if self.mysql_connect.if_insert_data(config.Student_data):
                QMessageBox.warning(self, '提示', '学号已经存在，请勿重复添加')
                return
            if self.mysql_connect.insert_data(config.Student_data):
                QMessageBox.information(self, '提示', '数据添加成功')
                self.home_1.show_student()
                return
        except:
            QMessageBox.warning(self, '错误', '数据库错误')


class home_alter(QWidget, Ui_home_alter):
    def __init__(self):
        super(home_alter, self).__init__()
        self.setupUi(self)
        self.home()

    def home(self):
        self.pushButton_add.clicked.connect(self.alter_datare)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def alter_datare(self):
        config.re_data.clear()
        config.re_data.append(self.lineEdit_name.text().strip())
        config.re_data.append(self.lineEdit_class.text().strip())
        config.re_data.append(self.lineEdit_yuwen.text().strip())
        config.re_data.append(self.lineEdit_math.text().strip())
        config.re_data.append(self.lineEdit_english.text().strip())
        config.re_data.append(config.row_data[0])
        print(config.re_data)
        self.mysql_connect = db.db_mysql()
        if self.mysql_connect.alter_data(config.re_data):
            QMessageBox.information(self, '提示', '信息修改成功')
            return
        QMessageBox.critical(self, '提示', '修改失败')


class query_data(QWidget, Ui_Form_Query):
    def __init__(self):
        super(query_data, self).__init__()
        self.setupUi(self)
        self.main()

    def main(self):
        self.pushButton.clicked.connect(self.Query_data)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def Query_data(self):
        sid = self.lineEdit_id.text().strip()
        name = self.lineEdit_name.text().strip()
        class_ = self.lineEdit_class.text().title()
        yuwen = self.lineEdit_yuwen.text().strip()
        math = self.lineEdit_math.text().strip()
        english = self.lineEdit_english.text().strip()
        if self.Query_data_vague(sid, name, class_, yuwen, math, english):
            print(config.result_data)
            self.tableWidget.setRowCount(0)
            self.tableWidget.clearContents()

            self.tableWidget.setColumnCount(6)
            rowcount = self.tableWidget.rowCount()
            for s in config.result_data:
                self.tableWidget.insertRow(rowcount)
                col = 0
                for key, value in s.items():
                    item = QTableWidgetItem(value)
                    self.tableWidget.setItem(rowcount, col, item)
                    col += 1
            return

    def Query_data_vague(self, *args):
        print(args)
        self.mysql_connect = db.db_mysql()
        if self.mysql_connect.query_data_vague(args):
            QMessageBox.information(self, '提示', '找到数据')
            return True
        QMessageBox.warning(self, '提示', '未找到数据')


class regedit(QWidget, Ui_Form_regedit):
    def __init__(self):
        super(regedit, self).__init__()
        self.setupUi(self)
        self.home()

    def home(self):
        self.pushButton_put.clicked.connect(self.register)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def register(self):
        self.mysql_connect = db.db_mysql()
        try:
            self.username = self.lineEdit_user.text().strip()
            self.password = self.lineEdit_pass.text().strip()
            self.repassword = self.lineEdit_repass.text().strip()
            if not self.username or not self.password:
                QMessageBox.warning(self, '提示', '用户名或密码不允许为空')
                return
            if self.mysql_connect.query_name(self.username):
                QMessageBox.information(self, '提示', '用户名{}已经注册了'.format(self.username))
                return
            if self.password == self.repassword:
                if self.mysql_connect.sql_users(self.username, self.password):
                    QMessageBox.information(self, '提示', '注册成功')
                else:
                    QMessageBox.critical(self, '提示', '未知错误')
            else:
                QMessageBox.warning(self, '提示', '两次输入的密码不相同')
        except:
            QMessageBox.warning(self, '提示', '数据库连接错误')


class repass(QWidget, Ui_Form_repass):
    def __init__(self):
        super(repass, self).__init__()
        self.setupUi(self)
        self.home()

    def home(self):
        self.lineEdit_user.setFocusPolicy(Qt.NoFocus)
        self.lineEdit_user.setText(config.login_username)
        self.pushButton_put.clicked.connect(self.repass)
        self.setWindowIcon(QIcon('./ui/icon.png'))

    def repass(self):
        self.username = self.lineEdit_user.text().strip()
        self.password = self.lineEdit_pass.text().strip()
        self.repassword = self.lineEdit_repass.text().strip()
        self.mysql_connect = db.db_mysql()
        if self.password == self.repassword:
            if self.mysql_connect.repass(self.password, config.login_username):
                QMessageBox.information(self, '提示', '{}密码修改成功'.format(config.login_username))
                return
            else:
                QMessageBox.warning(self, '提示', '密码修改失败')
                return
        QMessageBox.warning(self, '提示', '两次输入的密码不相同')


class del_class(QWidget, Ui_Form_del_user):
    def __init__(self):
        super(del_class, self).__init__()
        self.setupUi(self)
        self.home()

    def home(self):
        self.pushButton_put.clicked.connect(self.del_user)

    def del_user(self):
        self.mysql_connect = db.db_mysql()
        username = self.lineEdit_user.text()
        if username == config.login_username:
            QMessageBox.warning(self, '提示', '不能删除当前登录的账户')
            return
        if self.mysql_connect.query_name(username):
            choice = QMessageBox.question(self, '确定', '确定要删除{}用户吗'.format(username))
            if QMessageBox.Yes:
                if self.mysql_connect.del_user(username):
                    QMessageBox.information(self, '提示', '删除用户成功')
                    return
                QMessageBox.warning(self, '提示', '删除失败')
                return
            if QMessageBox.No:
                return
        QMessageBox.warning(self, '提示', '当前用户不存在')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # win = Winmain()
    # win.show()
    my = MyWindow()
    my.show()
    sys.exit(app.exec_())
